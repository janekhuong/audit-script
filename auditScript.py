import os
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import tabula
import re
from datetime import datetime


def pdf_to_csv(data):
    _, data_split = os.path.split(data)
    output = data_split.replace(".pdf", ".csv")

    tabula.convert_into(data, output, output_format="csv", pages="all", stream=True)

    return output


def fillExecs(data, excel):
    wb = load_workbook(excel)
    exec_sheet = wb[wb.sheetnames[2]]

    data_list = []

    with open(data, mode="r", encoding="utf-8") as file:
        for i, line in enumerate(file):
            if i == 0:
                continue
            row = line.strip().split(",")
            data_list.append(row)

    for row in range(6, len(data_list) + 6):
        i = 0
        for col in range(2, len(data_list[0]) + 2):
            exec_sheet[get_column_letter(col) + str(row)].value = data_list[row - 6][i]
            i += 1

    wb.save(excel)


def fillTransactions(data, excel):
    try:
        wb = load_workbook(excel)
    except FileNotFoundError:
        raise FileNotFoundError("The file was not found.")

    transaction_sheet = wb[wb.sheetnames[0]]

    data_list = []
    desired_columns = ["Customer Name", "Status", "Due Date", "Amount"]

    with open(data, mode="r", encoding="utf-8") as file:
        file.readline()
        header = file.readline().strip().split(",")
        column_indices = [header.index(col) for col in desired_columns]

        for line in file:
            fields = re.split(r',(?=(?:[^"]*"[^"]*")*[^"]*$)', line.strip())
            parsed_row = [fields[index] for index in column_indices]
            data_list.append(parsed_row)

    data_list = [row for row in data_list if row[1] == "Completed"]
    sorted_dates = sorted(
        [datetime.strptime(date[2].strip('"'), "%b %d, %Y") for date in data_list]
    )
    sorted_dates_string = [date.strftime("%b %d, %Y") for date in sorted_dates]

    sorted_data = []
    for date in sorted_dates_string:
        for row in data_list:
            if row[2].strip('"') == date:
                row[2] = datetime.strptime(row[2].strip('"'), "%b %d, %Y").strftime(
                    "%m/%d/%Y"
                )
                row[2], row[3] = row[3], row[2]
                row[2] = float(row[2])
                del row[1]
                sorted_data.append(row)
                data_list.remove(row)
                break

    for row in range(9, len(sorted_data) + 9):
        i = 0
        for col in [3, 10, 11]:
            transaction_sheet[get_column_letter(col) + str(row)].value = sorted_data[
                row - 9
            ][i]
            i += 1

        transaction_sheet["B" + str(row)].value = "Expense"
        transaction_sheet["G" + str(row)].value = "ACH"
        transaction_sheet["H" + str(row)].value = sorted_data[row - 9][0]

    wb.save(excel)


def main(command, data, excel):

    if not data.endswith(".csv"):
        if not data.endswith(".pdf"):
            raise ValueError("Data file must be a PDF")
        data = pdf_to_csv(data)

    if command == "fillExecs":
        fillExecs(data, excel)
    elif command == "fillTransactions":
        fillTransactions(data, excel)
    else:
        raise ValueError("Invalid command.")


if __name__ == "__main__":
    args = sys.argv
    if len(args) != 4:
        raise Exception("You must pass a command, a data file, and an xlsx file.")

    command, data, excel = args[1:]
    main(command, data, excel)
